from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\D9352\OneDrive - 基恩士（中国）有限公司\IT-PartnerShare - 文档\71. AD日志\AD_Log_Reports_Aggregated.xlsx')
ws = wb['事件详情']

print('检查"是否需要监测"列的所有非空值及其颜色:')
count = 0
for idx in range(2, ws.max_row + 1):
    cell = ws.cell(idx, 11)
    value = cell.value
    if value is not None:
        color = cell.fill.fgColor.rgb if cell.fill.fgColor else 'None'
        print(f'行{idx}: 值="{value}" 颜色={color}')
        count += 1
        if count >= 20:  # 只显示前20个非空值
            print('... (更多数据省略)')
            break

print(f'\n总共找到 {count} 个非空值')
