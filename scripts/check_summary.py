import pandas as pd
from openpyxl import load_workbook

file = r'C:\Users\D9352\OneDrive - 基恩士（中国）有限公司\IT-PartnerShare - 文档\71. AD日志\AD_Log_Reports_Aggregated_20260702_141704.xlsx'

# Read with openpyxl to see actual content
wb = load_workbook(file)
ws = wb['总体分析']

print('总体分析 sheet content:')
print('=' * 80)
for row in ws.iter_rows(values_only=True):
    if row[0]:
        print(row[0])
print('=' * 80)

# Check EventDetails for Critical/Error/Warning
events = pd.read_excel(file, sheet_name='EventDetails')
critical_error_warning = events[events['Level'].isin(['Critical', 'Error', 'Warning'])]

print(f'\nTotal Critical/Error/Warning events: {len(critical_error_warning)}')
print('\nSample with Event IDs:')
for idx, row in critical_error_warning.head(10).iterrows():
    eid = row['Event ID']
    level = row['Level']
    cause = row.get('可能原因', 'N/A')
    monitor = row.get('是否需要监测', 'N/A')
    cause_str = str(cause)[:50] if pd.notna(cause) else 'N/A'
    print(f'  EventID {eid} ({level}): {cause_str}...')
