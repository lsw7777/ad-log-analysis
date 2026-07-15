#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AD日志全功能聚合分析工具 - 全能版
整合：聚合统计 + AI智能分析 + 颜色高亮 + 智能修复
功能：
1. 聚合AD_Log_Reports目录下所有_stats.xlsx文件
2. 自动生成AI风险分析、服务器评估、执行摘要
3. 自动应用颜色高亮区分日志类型/服务器/级别
4. 智能修复可能原因与监测建议
输出：带时间戳的完整分析报告Excel
"""

import os
import glob
import shutil
import tempfile
import re
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# 全局常量定义（来自所有子脚本）
# ==============================================================================
EVENT_ANALYSIS = {
    4624: {"name": "登录成功", "category": "认证", "risk": "low",
           "desc": "成功登录事件。表示用户或计算机账户成功通过身份验证。"},
    4625: {"name": "登录失败", "category": "认证", "risk": "high",
           "desc": "登录失败事件。频繁的登录失败可能表示暴力破解攻击或凭据错误。"},
    4634: {"name": "注销", "category": "认证", "risk": "info",
           "desc": "用户注销事件，表示会话结束。"},
    4647: {"name": "用户发起注销", "category": "认证", "risk": "info",
           "desc": "用户主动注销，正常行为。"},
    4648: {"name": "显式凭据登录", "category": "认证", "risk": "medium",
           "desc": "使用显式凭据尝试登录，可能表示横向移动或使用存储凭据。"},
    4662: {"name": "AD对象访问", "category": "目录服务", "risk": "medium",
           "desc": "对Active Directory对象执行了操作。大量此事件可能表示目录枚举或权限探测。"},
    4663: {"name": "对象访问尝试", "category": "目录服务", "risk": "medium",
           "desc": "尝试访问对象，需关注是否涉及敏感AD对象。"},
    4670: {"name": "权限变更", "category": "权限管理", "risk": "high",
           "desc": "权限列表已更改。可能表示ACL修改，需确认是否为授权操作。"},
    4688: {"name": "进程创建", "category": "进程监控", "risk": "medium",
           "desc": "新进程已创建。关注是否创建了可疑工具（如cmd、powershell、mimikatz等）。"},
    4697: {"name": "服务安装", "category": "系统变更", "risk": "high",
           "desc": "系统上安装了新服务。恶意软件常通过安装服务实现持久化。"},
    4698: {"name": "计划任务创建", "category": "系统变更", "risk": "high",
           "desc": "创建了新的计划任务。攻击者常利用计划任务实现持久化和自动执行。"},
    4702: {"name": "计划任务更新", "category": "系统变更", "risk": "medium",
           "desc": "计划任务已更新。需确认修改内容是否合法。"},
    4719: {"name": "审计策略变更", "category": "安全策略", "risk": "critical",
           "desc": "系统审计策略已更改。攻击者可能关闭审计以掩盖活动。"},
    4720: {"name": "用户账户创建", "category": "账户管理", "risk": "high",
           "desc": "创建了新的用户账户。需确认是否为授权操作。"},
    4722: {"name": "用户账户启用", "category": "账户管理", "risk": "medium",
           "desc": "用户账户已启用。可能是重新激活已禁用账户。"},
    4724: {"name": "密码重置", "category": "账户管理", "risk": "high",
           "desc": "尝试重置用户密码。需确认是否为授权操作或社工攻击。"},
    4725: {"name": "用户账户禁用", "category": "账户管理", "risk": "medium",
           "desc": "用户账户被禁用。"},
    4726: {"name": "用户账户删除", "category": "账户管理", "risk": "high",
           "desc": "用户账户已删除。需确认是否为授权操作。"},
    4728: {"name": "成员添加到安全组", "category": "权限管理", "risk": "high",
           "desc": "成员被添加到安全组。如果是添加到Domain Admins等特权组，风险极高。"},
    4732: {"name": "成员添加到安全组", "category": "权限管理", "risk": "high",
           "desc": "成员被添加到安全启用组。关注是否被添加到管理员组。"},
    4738: {"name": "用户账户变更", "category": "账户管理", "risk": "medium",
           "desc": "用户账户属性已更改。需关注哪些属性被修改。"},
    4740: {"name": "账户锁定", "category": "认证", "risk": "medium",
           "desc": "账户因多次登录失败被锁定。可能表示暴力攻击或用户忘记密码。"},
    4741: {"name": "计算机账户创建", "category": "目录服务", "risk": "low",
           "desc": "新的计算机账户已创建。通常是新设备加入域的正常行为。"},
    4742: {"name": "计算机账户变更", "category": "目录服务", "risk": "low",
           "desc": "计算机账户属性已更改。"},
    4756: {"name": "成员添加到通用组", "category": "权限管理", "risk": "medium",
           "desc": "成员被添加到通用组。需确认目标组的权限级别。"},
    4767: {"name": "账户解锁", "category": "认证", "risk": "low",
           "desc": "账户已被解锁。"},
    4768: {"name": "Kerberos TGT请求", "category": "认证", "risk": "low",
           "desc": "Kerberos票据授予票据(TGT)请求。正常的域认证行为。"},
    4769: {"name": "Kerberos服务票据", "category": "认证", "risk": "low",
           "desc": "Kerberos服务票据请求。正常的域资源访问行为。"},
    4771: {"name": "Kerberos预认证失败", "category": "认证", "risk": "medium",
           "desc": "Kerberos预认证失败。可能是密码错误或配置问题。"},
    4776: {"name": "NTLM认证", "category": "认证", "risk": "medium",
           "desc": "使用了NTLM认证协议。NTLM较旧且安全性较低，建议迁移到Kerberos。"},
    4778: {"name": "RDP会话重连", "category": "远程访问", "risk": "low",
           "desc": "RDP远程桌面会话重新连接。"},
    4798: {"name": "本地组成员枚举", "category": "权限管理", "risk": "low",
           "desc": "枚举了本地组成员。"},
    5136: {"name": "目录服务对象修改", "category": "目录服务", "risk": "medium",
           "desc": "AD目录服务对象属性被修改。需关注修改了哪些属性。"},
    5137: {"name": "目录服务对象创建", "category": "目录服务", "risk": "medium",
           "desc": "创建了新的AD目录服务对象。"},
    5139: {"name": "目录服务对象移动", "category": "目录服务", "risk": "low",
           "desc": "AD对象在OU之间移动。"},
    5140: {"name": "网络共享访问", "category": "文件访问", "risk": "low",
           "desc": "访问了网络共享。"},
    5145: {"name": "共享对象权限检查", "category": "文件访问", "risk": "low",
           "desc": "检查了共享对象的权限。"},
    1102: {"name": "审计日志清除", "category": "安全策略", "risk": "critical",
           "desc": "安全审计日志已被清除！这是极高风险事件，可能表示攻击者试图掩盖痕迹。"},
    41: {"name": "意外关机", "category": "系统稳定性", "risk": "high",
         "desc": "系统意外断电或崩溃。可能导致数据丢失或系统损坏。"},
    1074: {"name": "系统关机/重启", "category": "系统管理", "risk": "low",
           "desc": "系统执行了关机或重启操作。"},
    6005: {"name": "事件日志服务启动", "category": "系统启动", "risk": "info",
           "desc": "事件日志服务已启动，系统开始记录日志。"},
    6006: {"name": "事件日志服务停止", "category": "系统关闭", "risk": "info",
           "desc": "事件日志服务已停止，系统正在关闭。"},
    6008: {"name": "意外关机记录", "category": "系统稳定性", "risk": "high",
           "desc": "记录了上一次意外关机事件。"},
    7031: {"name": "服务意外终止", "category": "服务状态", "risk": "medium",
           "desc": "服务意外终止。可能影响系统功能。"},
    7034: {"name": "服务意外崩溃", "category": "服务状态", "risk": "medium",
           "desc": "服务意外崩溃。需检查服务配置和依赖项。"},
    7036: {"name": "服务状态变更", "category": "服务状态", "risk": "info",
           "desc": "服务启动或停止。"},
    7040: {"name": "服务启动类型变更", "category": "服务状态", "risk": "medium",
           "desc": "服务的启动类型已更改。"},
    1001: {"name": "Windows错误报告", "category": "系统错误", "risk": "medium",
           "desc": "Windows报告了一个错误。需关注BSOD等严重错误。"},
    2001: {"name": "驱动程序加载", "category": "系统", "risk": "medium",
           "desc": "加载了新的驱动程序。"},
}

RISK_COLORS = {
    "critical": "FF0000",
    "high": "FF6600",
    "medium": "FFCC00",
    "low": "00CC00",
    "info": "999999",
}

RISK_LABELS = {
    "critical": "严重",
    "high": "高",
    "medium": "中",
    "low": "低",
    "info": "信息",
}

SUSPICIOUS_PROCESSES = [
    "mimikatz", "procdump", "lsass", "ntdsutil", "secretsdump",
    "powershell -enc", "powershell -e ", "invoke-", "downloadstring",
    "net user", "net group", "net localgroup", "whoami", "systeminfo",
    "ipconfig", "netstat", "tasklist", "reg save", "reg export",
    "wmic", "vssadmin", "bcdedit", "schtasks", "sc create",
]

COLORS = {
    'log_type': {
        '安全': 'D6EAF8',
        '系统': 'D5F5E3',
        '应用程序': 'FDEBD0',
        '设置': 'E8DAEF',
    },
    'ad_server': {
        '10.59.91.1': 'FADBD8',
        '10.59.91.2': 'D1F2EB',
        '10.59.97.1': 'FCF3CF',
        '10.59.98.1': 'D6EAF8',
        '10.59.99.1': 'E8DAEF',
    },
    'level': {
        'Critical': 'FF0000',
        'Error': 'FF6600',
        'Warning': 'FFCC00',
        'Information': '92D050',
        'Log': '00B0F0',
        'Verbose': 'BDD7EE',
    },
    'monitor': {
        '是': 'FF9999',
        '否': 'C6EFCE',
        '视情况': 'FFEB9C',
    },
}

# ==============================================================================
# 模块1：聚合统计（来自 aggregate_stats.py）
# ==============================================================================
def aggregate_stats(root_dir):
    reports_dir = os.path.join(root_dir, "AD_Log_Reports")
    pattern = os.path.join(reports_dir, "*_stats.xlsx")
    excel_files = sorted(glob.glob(pattern))
    
    if not excel_files:
        print("未找到任何*_stats.xlsx文件")
        return None, None
    
    print(f"找到 {len(excel_files)} 个统计文件")
    tmp_dir = tempfile.mkdtemp(prefix="ad_agg_")
    all_summary = []
    all_event_details = []
    
    for file_path in excel_files:
        file_name = os.path.basename(file_path)
        print(f"处理: {file_name}")
        tmp_file = os.path.join(tmp_dir, file_name)
        try:
            shutil.copy2(file_path, tmp_file)
        except Exception as e:
            print(f"  复制失败: {e}")
            continue
        
        try:
            summary_df = pd.read_excel(tmp_file, sheet_name="Summary")
            base_name = file_name.replace("_stats.xlsx", "")
            parts = base_name.rsplit("_", 1)
            ip_address = parts[0] if len(parts) == 2 else base_name
            log_type = parts[1] if len(parts) == 2 else ""
            
            summary_df.insert(0, "SourceFile", file_name)
            summary_df.insert(1, "IPAddress", ip_address)
            summary_df.insert(2, "LogType", log_type)
            all_summary.append(summary_df)
            
            try:
                event_df = pd.read_excel(tmp_file, sheet_name="EventDetails")
            except ValueError:
                try:
                    event_df = pd.read_excel(tmp_file, sheet_name="事件详情")
                except ValueError:
                    event_df = pd.read_excel(tmp_file, sheet_name="事件分组")
            
            event_df.insert(0, "SourceFile", file_name)
            event_df.insert(1, "IPAddress", ip_address)
            event_df.insert(2, "LogType", log_type)
            all_event_details.append(event_df)
            
            print(f"  成功: Summary {len(summary_df)} 行, EventDetails {len(event_df)} 行")
        except Exception as e:
            print(f"  读取错误: {e}")
            continue
    
    shutil.rmtree(tmp_dir, ignore_errors=True)
    if not all_summary:
        print("没有成功读取任何数据")
        return None, None
    
    combined_summary = pd.concat(all_summary, ignore_index=True)
    combined_event_details = pd.concat(all_event_details, ignore_index=True)
    return combined_summary, combined_event_details

# ==============================================================================
# 模块2：AI分析核心函数（来自 ai_analysis.py）
# ==============================================================================
def analyze_event(row):
    eid = row.get("Event ID", 0)
    provider = str(row.get("Provider Name", ""))
    level = str(row.get("Level", ""))
    count = int(row.get("Count", 0))
    msg = str(row.get("Sample Message", ""))
    log_type = str(row.get("LogType", ""))
    first_seen = str(row.get("First Seen", ""))
    last_seen = str(row.get("Last Seen", ""))

    info = EVENT_ANALYSIS.get(eid, None)
    if info:
        event_name = info["name"]
        category = info["category"]
        base_risk = info["risk"]
        desc = info["desc"]
    else:
        event_name = f"事件ID {eid}"
        category = "其他"
        base_risk = "info"
        desc = f"未分类事件ID {eid}，来自{provider}。"

    risk = base_risk
    extras = []

    if eid == 4625 and count > 100:
        risk = "critical"
        extras.append(f"⚠ 失败次数高达{count}次，极可能遭遇暴力破解攻击")
    elif eid == 4625 and count > 20:
        risk = "high"
        extras.append(f"⚠ 失败{count}次，可能存在密码猜测行为")

    if eid == 4740 and count > 5:
        risk = "high"
        extras.append(f"⚠ {count}个账户被锁定，可能存在大规模攻击")

    if eid == 1102:
        risk = "critical"
        extras.append("🚨 审计日志被清除，高度可疑！")

    if eid == 4688:
        msg_lower = msg.lower()
        for sp in SUSPICIOUS_PROCESSES:
            if sp in msg_lower:
                risk = "critical"
                extras.append(f"🚨 检测到可疑进程关键词: {sp}")
                break

    if eid in (4720, 4726, 4724, 4728, 4732) and count > 5:
        if risk in ("low", "info"):
            risk = "high"
        extras.append(f"⚠ 发生{count}次，需确认是否为批量操作")

    if eid in (4698, 4697):
        if "powershell" in msg.lower() or "cmd" in msg.lower():
            risk = "critical"
            extras.append("🚨 计划任务/服务涉及命令行执行，可能为持久化后门")

    if count > 10000:
        extras.append(f"📊 事件量极大({count}次)，建议检查是否存在异常循环")
    elif count > 1000:
        extras.append(f"📊 事件量较大({count}次)")

    if first_seen and last_seen and first_seen != last_seen:
        try:
            t1 = pd.to_datetime(first_seen)
            t2 = pd.to_datetime(last_seen)
            span = t2 - t1
            if span.total_seconds() > 0:
                extras.append(f"⏱ 时间跨度: {span}")
        except:
            pass

    analysis_parts = [f"[{event_name}]", f"分类: {category}", f"风险: {RISK_LABELS.get(risk, risk)}"]
    if extras:
        analysis_parts.extend(extras)
    else:
        analysis_parts.append(desc)

    return risk, " | ".join(analysis_parts)

def generate_server_analysis(server_summary, df_events):
    lines = []
    ip = server_summary["IPAddress"].iloc[0] if len(server_summary) > 0 else "未知"
    lines.append(f"📋 服务器 {ip} 综合分析")
    lines.append("=" * 50)

    total_events = 0
    total_critical = 0
    total_error = 0
    total_warning = 0

    for _, row in server_summary.iterrows():
        level_col = row.get("Unnamed: 0", "")
        count_col = row.get("Unnamed: 1", 0)
        level = str(level_col).strip()
        try:
            count = int(count_col)
        except:
            count = 0
        if level == "Total Events":
            total_events += count
        elif level == "Critical":
            total_critical += count
        elif level == "Error":
            total_error += count
        elif level == "Warning":
            total_warning += count

    lines.append(f"总事件数: {total_events:,}")
    lines.append(f"严重事件: {total_critical:,}")
    lines.append(f"错误事件: {total_error:,}")
    lines.append(f"警告事件: {total_warning:,}")
    lines.append("")

    server_events = df_events[df_events["IPAddress"] == ip]
    critical_events = server_events[server_events["Risk Level"] == "critical"]
    high_events = server_events[server_events["Risk Level"] == "high"]
    
    if len(critical_events) > 0:
        lines.append("🚨 严重风险事件:")
        for _, row in critical_events.iterrows():
            lines.append(f"  - EventID {row['Event ID']} ({row['Provider Name']}): {row['Count']}次")
        lines.append("")
    
    if len(high_events) > 0:
        lines.append("⚠ 高风险事件:")
        for _, row in high_events.head(10).iterrows():
            lines.append(f"  - EventID {row['Event ID']} ({row['Provider Name']}): {row['Count']}次")
        lines.append("")

    if total_critical > 0:
        lines.append("🔴 风险等级: 严重 - 存在严重级别事件，建议立即排查")
    elif len(high_events) > 0:
        lines.append("🟠 风险等级: 高 - 存在高风险事件，需要关注")
    elif total_error > 100:
        lines.append("🟡 风险等级: 中 - 错误事件较多，建议检查系统健康状态")
    else:
        lines.append("🟢 风险等级: 正常")
    lines.append("")
    return "\n".join(lines)

def generate_risk_assessment(df_events):
    lines = []
    lines.append("🔍 AD日志总体风险评估报告")
    lines.append("=" * 60)
    lines.append(f"分析时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"分析服务器数: {df_events['IPAddress'].nunique()}")
    lines.append(f"总事件记录数: {len(df_events):,}")
    lines.append("")

    risk_counts = df_events["Risk Level"].value_counts()
    lines.append("【风险等级分布】")
    for r in ["critical", "high", "medium", "low", "info"]:
        c = risk_counts.get(r, 0)
        lines.append(f"  {RISK_LABELS.get(r, r)}: {c} 条记录")
    lines.append("")

    critical_events = df_events[df_events["Risk Level"] == "critical"]
    if len(critical_events) > 0:
        lines.append("【严重风险事件】")
        for _, row in critical_events.iterrows():
            lines.append(f"  [{row['IPAddress']}] {row['LogType']} - EventID {row['Event ID']} ({row['Provider Name']}): {row['Count']}次")
        lines.append("")

    high_events = df_events[df_events["Risk Level"] == "high"]
    if len(high_events) > 0:
        lines.append("【高风险事件】")
        for _, row in high_events.head(20).iterrows():
            lines.append(f"  [{row['IPAddress']}] {row['LogType']} - EventID {row['Event ID']} ({row['Provider Name']}): {row['Count']}次")
        lines.append("")

    servers = df_events.groupby("IPAddress")
    lines.append("【各服务器风险概要】")
    for ip, group in servers:
        c_count = len(group[group["Risk Level"] == "critical"])
        h_count = len(group[group["Risk Level"] == "high"])
        total_count = group["Count"].sum()
        if c_count > 0:
            risk_level = "🔴 严重"
        elif h_count > 0:
            risk_level = "🟠 高"
        else:
            risk_level = "🟢 正常"
        lines.append(f"  {ip}: 风险等级 {risk_level} | 严重:{c_count} 高:{h_count} | 总事件:{total_count:,}")
    lines.append("")

    logon_failures = df_events[df_events["Event ID"] == 4625]
    if len(logon_failures) > 0:
        lines.append("【登录失败分析】")
        total_failures = logon_failures["Count"].sum()
        lines.append(f"  总登录失败次数: {total_failures:,}")
        for _, row in logon_failures.iterrows():
            lines.append(f"  [{row['IPAddress']}] {row['Count']}次失败")
        if total_failures > 1000:
            lines.append("  ⚠ 登录失败次数极高，强烈建议检查是否存在暴力破解")
        lines.append("")

    account_events = df_events[df_events["Event ID"].isin([4720, 4722, 4724, 4725, 4726])]
    if len(account_events) > 0:
        lines.append("【账户变更分析】")
        for _, row in account_events.iterrows():
            eid_name = EVENT_ANALYSIS.get(row["Event ID"], {}).get("name", f"EventID {row['Event ID']}")
            lines.append(f"  [{row['IPAddress']}] {eid_name}: {row['Count']}次")
        lines.append("")

    lines.append("【建议措施】")
    if len(critical_events) > 0:
        lines.append("  1. 立即调查所有严重风险事件")
        lines.append("  2. 检查是否有未授权的账户创建/权限变更")
    if len(logon_failures) > 0 and logon_failures["Count"].sum() > 100:
        lines.append("  3. 审查账户锁定策略，考虑启用MFA")
    lines.append("  4. 定期检查AD审计策略是否完整启用")
    lines.append("  5. 确保所有AD服务器时间同步(NTP)")
    lines.append("  6. 备份当前日志并归档")
    return "\n".join(lines)

def generate_executive_summary(df_events, df_summary):
    lines = []
    lines.append("=" * 70)
    lines.append("📊 AD环境安全态势总体分析报告")
    lines.append("=" * 70)
    lines.append(f"报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"分析范围: {df_events['IPAddress'].nunique()} 台AD服务器")
    lines.append(f"日志时间跨度: 查看各服务器日志")
    lines.append("")
    
    critical_count = len(df_events[df_events["Risk Level"] == "critical"])
    high_count = len(df_events[df_events["Risk Level"] == "high"])
    medium_count = len(df_events[df_events["Risk Level"] == "medium"])
    
    if critical_count > 0:
        overall_risk = "🔴 高风险 (Critical)"
        risk_score = "严重"
    elif high_count > 5:
        overall_risk = "🟠 较高风险 (High)"
        risk_score = "高"
    elif high_count > 0 or medium_count > 10:
        overall_risk = "🟡 中等风险 (Medium)"
        risk_score = "中"
    else:
        overall_risk = "🟢 低风险 (Low)"
        risk_score = "低"
    
    lines.append(f"【总体风险评级】: {overall_risk}")
    lines.append(f"风险评分: {risk_score}")
    lines.append("")
    
    lines.append("【执行摘要】")
    lines.append(f"本次分析共检查了 {len(df_events):,} 条事件记录，发现以下关键问题：")
    lines.append("")
    
    finding_num = 1
    if critical_count > 0:
        critical_events = df_events[df_events["Risk Level"] == "critical"]
        lines.append(f"{finding_num}. 【严重】发现 {critical_count} 个严重风险事件")
        for _, row in critical_events.head(3).iterrows():
            lines.append(f"   - {row['IPAddress']}: EventID {row['Event ID']} ({row['Count']}次)")
        lines.append("")
        finding_num += 1
    
    login_failures = df_events[df_events["Event ID"] == 4625]
    if len(login_failures) > 0:
        total_failures = login_failures["Count"].sum()
        lines.append(f"{finding_num}. 【认证安全】登录失败总计 {total_failures:,} 次")
        if total_failures > 1000:
            lines.append("   ⚠ 失败次数极高，可能存在暴力破解攻击或账户配置问题")
            lines.append("   建议: 检查账户锁定策略，启用多因素认证(MFA)")
        elif total_failures > 100:
            lines.append("   ⚠ 失败次数较多，建议关注")
        for _, row in login_failures.head(5).iterrows():
            lines.append(f"   - {row['IPAddress']}: {row['Count']}次失败")
        lines.append("")
        finding_num += 1
    
    account_changes = df_events[df_events["Event ID"].isin([4720, 4722, 4724, 4725, 4726, 4728, 4732])]
    if len(account_changes) > 0:
        total_changes = account_changes["Count"].sum()
        lines.append(f"{finding_num}. 【账户管理】账户变更事件 {total_changes:,} 次")
        lines.append("   包括: 账户创建/删除/启用/禁用/密码重置/组成员变更")
        if total_changes > 100:
            lines.append("   ⚠ 变更频繁，建议审查是否为授权操作")
        lines.append("   建议: 建立账户变更审批流程，定期审计特权账户")
        lines.append("")
        finding_num += 1
    
    persistence_events = df_events[df_events["Event ID"].isin([4697, 4698, 4702])]
    if len(persistence_events) > 0:
        total_persistence = persistence_events["Count"].sum()
        lines.append(f"{finding_num}. 【持久化风险】计划任务/服务变更 {total_persistence:,} 次")
        lines.append("   攻击者常利用计划任务和服务实现持久化")
        lines.append("   建议: 审查新建的计划任务和服务，确认是否为合法软件")
        lines.append("")
        finding_num += 1
    
    system_issues = df_events[df_events["Event ID"].isin([41, 6008, 7031, 7034])]
    if len(system_issues) > 0:
        total_issues = system_issues["Count"].sum()
        lines.append(f"{finding_num}. 【系统稳定性】系统异常事件 {total_issues:,} 次")
        lines.append("   包括: 意外关机、服务崩溃等")
        lines.append("   建议: 检查硬件健康状态，查看系统日志定位根因")
        lines.append("")
        finding_num += 1
    
    lines.append("【各服务器详细分析】")
    lines.append("")
    for ip in sorted(df_events["IPAddress"].unique()):
        server_events = df_events[df_events["IPAddress"] == ip]
        server_critical = len(server_events[server_events["Risk Level"] == "critical"])
        server_high = len(server_events[server_events["Risk Level"] == "high"])
        server_total = server_events["Count"].sum()
        
        if server_critical > 0:
            status = "🔴 严重风险"
        elif server_high > 0:
            status = "🟠 高风险"
        elif len(server_events[server_events["Risk Level"] == "medium"]) > 0:
            status = "🟡 中等风险"
        else:
            status = "🟢 正常"
        
        lines.append(f"▶ {ip} - {status}")
        lines.append(f"  总事件: {server_total:,} | 严重: {server_critical} | 高: {server_high}")
        
        server_login_fail = server_events[server_events["Event ID"] == 4625]["Count"].sum()
        if server_login_fail > 100:
            lines.append(f"  ⚠ 登录失败 {server_login_fail:,} 次 - 可能存在暴力破解")
        
        server_account_changes = server_events[server_events["Event ID"].isin([4720, 4724, 4726])]["Count"].sum()
        if server_account_changes > 50:
            lines.append(f"  ⚠ 账户变更 {server_account_changes:,} 次 - 建议审查")
        
        server_persistence = server_events[server_events["Event ID"].isin([4697, 4698])]["Count"].sum()
        if server_persistence > 0:
            lines.append(f"  ⚠ 计划任务/服务创建 {server_persistence} 次 - 需确认合法性")
        lines.append("")
    
    lines.append("=" * 70)
    lines.append("【优先级建议】")
    lines.append("=" * 70)
    lines.append("")
    
    priority = 1
    if critical_count > 0:
        lines.append(f"【P0 - 立即处理】")
        lines.append(f"{priority}. 调查所有严重风险事件，确认是否为安全事件")
        lines.append("   - 检查是否有未授权的账户操作")
        lines.append("   - 检查是否有审计日志被清除")
        lines.append("")
        priority += 1
    
    if len(login_failures) > 0 and login_failures["Count"].sum() > 100:
        lines.append(f"【P1 - 高优先级】")
        lines.append(f"{priority}. 加强认证安全")
        lines.append("   - 审查账户锁定策略（建议5次失败锁定30分钟）")
        lines.append("   - 为特权账户启用多因素认证(MFA)")
        lines.append("   - 检查是否有来自异常IP的登录尝试")
        lines.append("")
        priority += 1
    
    if len(account_changes) > 0 and account_changes["Count"].sum() > 50:
        lines.append(f"【P2 - 中优先级】")
        lines.append(f"{priority}. 完善账户管理流程")
        lines.append("   - 建立账户变更审批和审计机制")
        lines.append("   - 定期审查特权组成员（Domain Admins等）")
        lines.append("   - 清理不再使用的账户")
        lines.append("")
        priority += 1
    
    if len(persistence_events) > 0:
        lines.append(f"【P3 - 定期审查】")
        lines.append(f"{priority}. 审查系统变更")
        lines.append("   - 审查新建的计划任务和服务")
        lines.append("   - 确认所有变更都有变更记录")
        lines.append("   - 使用白名单机制控制软件安装")
        lines.append("")
        priority += 1
    
    lines.append(f"【P4 - 持续改进】")
    lines.append(f"{priority}. 建立安全基线")
    lines.append("   - 定期备份和归档日志")
    lines.append("   - 确保所有AD服务器时间同步(NTP)")
    lines.append("   - 定期执行安全评估和渗透测试")
    lines.append("   - 建立安全事件响应流程")
    lines.append("")
    
    lines.append("=" * 70)
    lines.append("报告结束")
    lines.append("=" * 70)
    return "\n".join(lines)

# ==============================================================================
# 模块3：颜色样式（来自 apply_colors.py）
# ==============================================================================
def get_col_index(ws, name):
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == name:
            return idx
    return None

def apply_colors_to_wb(wb):
    try:
        ws = wb['事件详情']
    except:
        try:
            ws = wb['EventDetails']
        except:
            return
    
    col_log = get_col_index(ws, '日志类型') or get_col_index(ws, 'LogType')
    col_server = get_col_index(ws, 'AD服务器') or get_col_index(ws, 'IPAddress')
    col_level = get_col_index(ws, '级别') or get_col_index(ws, 'Level')
    col_monitor = get_col_index(ws, '是否需要监测') or get_col_index(ws, '是否监测')
    
    thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if col_log:
            val = row[col_log - 1].value
            if val and str(val) in COLORS['log_type']:
                row[col_log - 1].fill = PatternFill('solid', fgColor=COLORS['log_type'][str(val)])
        
        if col_server:
            val = row[col_server - 1].value
            if val and str(val) in COLORS['ad_server']:
                row[col_server - 1].fill = PatternFill('solid', fgColor=COLORS['ad_server'][str(val)])
        
        if col_level:
            val = row[col_level - 1].value
            if val and str(val) in COLORS['level']:
                color = COLORS['level'][str(val)]
                row[col_level - 1].fill = PatternFill('solid', fgColor=color)
                row[col_level - 1].font = Font(bold=True, color='FFFFFF' if str(val) in ('Critical',) else '000000')
        
        if col_monitor:
            val = row[col_monitor - 1].value
            if val:
                val_str = str(val)
                for key, color in COLORS['monitor'].items():
                    if val_str.startswith(key):
                        row[col_monitor - 1].fill = PatternFill('solid', fgColor=color)
                        break
    
    try:
        ws_sum = wb['汇总']
    except:
        try:
            ws_sum = wb['Summary']
        except:
            return
    
    for col_idx in range(1, ws_sum.max_column + 1):
        header = ws_sum.cell(1, col_idx).value
        if header in COLORS['level']:
            color = COLORS['level'][header]
            for row_idx in range(2, ws_sum.max_row + 1):
                cell = ws_sum.cell(row_idx, col_idx)
                cell.fill = PatternFill('solid', fgColor=color)
                cell.font = Font(bold=True, color='FFFFFF' if header == 'Critical' else '000000')

def style_excel(wb, df_events):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    risk_fills = {}
    for risk_key, color in RISK_COLORS.items():
        risk_fills[risk_key] = PatternFill(start_color=color, end_color=color, fill_type="solid")

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row_idx in range(1, min(ws.max_row + 1, 100)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, min(len(str(val)), 60))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    if "EventDetails" in wb.sheetnames:
        ws = wb["EventDetails"]
        risk_col_idx = None
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col_idx).value == "Risk Level":
                risk_col_idx = col_idx
                break
        if risk_col_idx:
            for row_idx in range(2, ws.max_row + 1):
                risk_val = str(ws.cell(row=row_idx, column=risk_col_idx).value or "").lower()
                if risk_val in risk_fills:
                    ws.cell(row=row_idx, column=risk_col_idx).fill = risk_fills[risk_val]
    
    apply_colors_to_wb(wb)

# ==============================================================================
# 模块4：智能修复（来自 smart_fix.py）
# ==============================================================================
def analyze_event_content(event_id, provider, message):
    if not message or message == 'nan':
        message = ""
    message_lower = message.lower()
    event_id = int(event_id)
    cause = ""
    monitor = ""

    if event_id in [4, 3] and "更新" in message:
        if "重启" in message or "restart" in message_lower:
            cause = "Windows更新已下载但需要重启才能完成安装。在重启之前，系统可能缺少最新的安全补丁。"
            monitor = "是 - 建议尽快安排重启以完成更新"
        elif "未完成" in message or "pending" in message_lower:
            cause = "系统有未完成的更新事务，之前的更新或系统变更没有彻底完成。"
            monitor = "是 - 建议检查Windows Update状态并重启系统"
        else:
            cause = "Windows Update相关事件，可能涉及更新安装或配置。"
            monitor = "视情况 - 检查Windows Update状态"
    elif event_id == 21 and ("智能卡" in message or "smart card" in message_lower or "证书" in message):
        cause = "智能卡登录失败，证书可能已过期、被吊销，或证书链不受信任。"
        monitor = "是 - 检查证书有效期和信任链"
    elif event_id == 9009 and ("IIS" in message or "web" in message_lower):
        cause = "IIS在清理历史配置目录时出错，不影响IIS正常运行。"
        monitor = "否 - 可以忽略，不影响服务"
    elif event_id == 10149 and ("WinRM" in message or "远程管理" in message):
        cause = "WinRM（Windows远程管理）服务未运行，无法通过PowerShell Remoting等工具远程管理。"
        monitor = "是 - 如需远程管理，应启动WinRM服务"
    elif event_id == 10016 and ("权限" in message or "DCOM" in message or "COM" in message):
        if "RuntimeBroker" in message:
            cause = "DCOM权限配置问题，RuntimeBroker访问被拒绝。微软官方表示通常可安全忽略。"
            monitor = "否 - 常见系统提示，可忽略"
        elif "CDP" in message or "Activity Store" in message:
            cause = "DCOM权限问题，系统组件访问被拒绝。通常不影响系统功能。"
            monitor = "否 - 可安全忽略"
        else:
            cause = "DCOM权限配置不完整，某个程序没有权限访问COM组件。"
            monitor = "视情况 - 如系统功能正常可忽略"
    elif event_id == 6040 and ("NTLM" in message or "认证" in message):
        cause = "NTLM认证请求被拒绝，目标名称不匹配。可能是配置错误或中间人攻击。"
        monitor = "是 - 检查NTLM配置和网络安全"
    elif event_id == 6038 and "NTLM" in message:
        cause = "客户端使用NTLM认证（而非更安全的Kerberos）。NTLM存在已知安全漏洞。"
        monitor = "是 - 建议迁移到Kerberos认证"
    elif event_id == 32 and ("磁盘" in message or "disk" in message_lower or "写入缓存" in message):
        cause = "磁盘启用了写入缓存，突然断电可能丢失缓存数据。"
        monitor = "否 - 正常配置，确保有UPS保护"
    elif event_id == 32 and ("KDC" in message or "证书" in message):
        cause = "域控制器KDC证书配置不正确，缺少KDC扩展密钥用法。未加入域的设备无法通过证书认证。"
        monitor = "是 - 修复KDC证书配置"
    elif event_id in [58, 158] and ("磁盘" in message or "签名" in message):
        cause = "多块磁盘使用相同标识符，系统可能混淆磁盘，导致数据读写错误。"
        monitor = "是 - 立即检查并修复磁盘签名冲突"
    elif event_id == 157 and ("磁盘" in message or "移除" in message):
        cause = "磁盘被意外移除，可能是硬盘故障、连接松动或存储链路中断。"
        monitor = "是 - 立即检查硬件状态"
    elif event_id == 153 and ("磁盘" in message or "IO" in message or "重试" in message):
        cause = "磁盘进行IO重试，可能表示磁盘有坏道或连接不稳定，是硬盘故障的早期信号。"
        monitor = "是 - 建议检查磁盘健康状态"
    elif event_id == 5840 and ("WAS" in message or "进程激活" in message):
        cause = "Windows进程激活服务以经典模式启动，IIS相关配置信息，通常不影响功能。"
        monitor = "否 - 正常配置信息"
    elif event_id == 7009 and ("启动" in message or "超时" in message):
        if "Zabbix" in message:
            cause = "Zabbix Agent启动超时，开机时资源竞争导致。"
            monitor = "否 - 常见开机现象，可忽略"
        else:
            cause = "服务启动超时，未在默认30秒内完成启动。"
            monitor = "是 - 检查服务状态和系统资源"
    elif event_id == 7000 and ("启动" in message or "失败" in message):
        cause = "服务启动失败，可能是账户权限不足或依赖服务未启动。"
        monitor = "是 - 检查服务配置和依赖关系"
    elif event_id == 10029 and ("TrustedInstaller" in message or "等待" in message):
        cause = "系统在等待TrustedInstaller停止时超时，另一个组件因等待超时而激活失败。"
        monitor = "否 - 偶发事件，可忽略"
    elif event_id == 5723 and ("信任" in message or "安全通道" in message or "Netlogon" in message):
        if "找不到" in message or "not found" in message_lower:
            cause = "域控制器安全数据库中找不到计算机账户，信任关系断裂。"
            monitor = "是 - 立即检查计算机账户和AD复制"
        else:
            cause = "计算机与域控制器之间的信任关系或安全通道建立失败。"
            monitor = "是 - 检查网络连接和域信任关系"
    elif event_id == 5722 and ("信任" in message or "Kerberos" in message or "Netlogon" in message):
        cause = "计算机账户与域控制器之间的信任关系或Kerberos身份验证失效。"
        monitor = "是 - 检查网络连接和域控制器状态"
    elif event_id == 5805 and ("拒绝" in message or "访问" in message or "Netlogon" in message):
        if "信任关系" in message or "断裂" in message:
            cause = "计算机与域之间的信任关系已彻底断裂，身份验证被拒绝。"
            monitor = "是 - 立即修复计算机信任关系"
        else:
            cause = "身份验证尝试被目标服务器拒绝，可能是安全通道问题。"
            monitor = "是 - 检查网络连接和信任关系"
    elif event_id == 5838 and ("Netlogon" in message or "RPC" in message or "签名" in message):
        cause = "客户端使用旧版本Netlogon协议，安全级别低于域控制器要求。"
        monitor = "是 - 升级客户端Netlogon协议版本"
    elif event_id == 5807 and ("站点" in message or "AD站点" in message):
        cause = "客户端连接到没有映射到AD站点的域控，AD站点配置不完整，可能导致登录变慢。"
        monitor = "是 - 完善AD站点和子网配置"
    elif event_id == 5827 and ("Netlogon" in message or "安全" in message):
        cause = "Netlogon安全策略升级后，旧协议被拒绝。"
        monitor = "是 - 检查并升级客户端安全协议"
    elif event_id == 14 and ("Kerberos" in message or "加密" in message):
        cause = "Kerberos加密类型不匹配，请求的加密类型与账户可用加密类型不一致。"
        monitor = "是 - 检查Kerberos加密配置"
    elif event_id == 30 and ("ETW" in message or "权限" in message):
        cause = "ETW提供程序启用了安全机制，但事件日志服务账户缺少必要权限。"
        monitor = "是 - 修复ETW提供程序权限配置"
    elif event_id == 1077 and ("注销" in message or "失败" in message):
        cause = "注销操作失败，可能有程序阻止了注销。"
        monitor = "否 - 偶发事件，可忽略"
    elif event_id