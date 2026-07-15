import shutil, tempfile, os
from openpyxl import load_workbook

src = r'C:\Users\D9352\OneDrive - 基恩士（中国）有限公司\IT-PartnerShare - 文档\71. AD日志\AD_Log_Reports_Aggregated.xlsx'
tmp = os.path.join(tempfile.gettempdir(), 'check_monitor_colors.xlsx')
shutil.copy2(src, tmp)

wb = load_workbook(tmp)
ws = wb['事件详情']

# 找到"是否监测"列
monitor_col = None
for col_idx in range(1, ws.max_column + 1):
    if ws.cell(1, col_idx).value == '是否监测':
        monitor_col = col_idx
        break

print(f'"是否监测"列索引: {monitor_col}')

# 检查前20行的颜色
for row_idx in range(2, min(22, ws.max_row + 1)):
    cell = ws.cell(row_idx, monitor_col)
    value = cell.value
    fill_color = cell.fill.fgColor.rgb if cell.fill.fgColor else 'None'
    print(f'行{row_idx}: 值="{value}" 颜色={fill_color}')
