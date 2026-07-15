#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os
import shutil
import tempfile
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Event ID analysis database
EVENT_CAUSES = {
    # Critical events
    1102: {
        "cause": "审计日志被清除，可能是攻击者试图掩盖痕迹或管理员误操作",
        "monitor": "是 - 立即调查清除日志的账户和时间"
    },
    4719: {
        "cause": "系统审计策略被更改，可能影响安全监控能力",
        "monitor": "是 - 确认是否为授权变更"
    },
    
    # System events
    4: {
        "cause": "审计策略变更事件，系统安全策略被修改",
        "monitor": "是 - 确认是否为授权变更"
    },
    13: {
        "cause": "Windows时间服务事件，可能是时间同步问题",
        "monitor": "否 - 通常是正常时间同步"
    },
    32: {
        "cause": "事件日志服务启动，系统开始记录日志",
        "monitor": "否 - 正常系统启动"
    },
    
    # Application and service errors
    1000: {
        "cause": "应用程序崩溃或错误，可能是软件缺陷或配置问题",
        "monitor": "是 - 检查应用程序日志定位根因"
    },
    10010: {
        "cause": "服务控制管理器错误，服务启动或停止失败",
        "monitor": "是 - 检查服务状态和依赖关系"
    },
    10016: {
        "cause": "DCOM错误，分布式组件对象模型通信失败",
        "monitor": "视情况 - 可能是权限或网络问题"
    },
    
    # Audit and security events
    5612: {
        "cause": "审计策略已修改，安全审计配置变更",
        "monitor": "是 - 确认是否为授权变更"
    },
    
    # Netlogon errors
    5722: {
        "cause": "Netlogon错误，域控制器通信失败或认证问题",
        "monitor": "是 - 检查网络连接和域控制器状态"
    },
    5723: {
        "cause": "Netlogon错误，安全通道建立失败",
        "monitor": "是 - 检查域信任关系和网络连接"
    },
    
    # System events
    5805: {
        "cause": "系统事件，可能是硬件或驱动程序问题",
        "monitor": "视情况 - 检查系统日志"
    },
    5807: {
        "cause": "系统事件，可能是配置或权限问题",
        "monitor": "视情况 - 检查系统日志"
    },
    5838: {
        "cause": "系统事件，可能是服务或进程问题",
        "monitor": "视情况 - 检查系统日志"
    },
    5840: {
        "cause": "系统事件，可能是资源或性能问题",
        "monitor": "视情况 - 检查系统日志"
    },
    
    # Event log service
    6005: {
        "cause": "事件日志服务启动，系统开始记录日志",
        "monitor": "否 - 正常系统启动"
    },
    6006: {
        "cause": "事件日志服务停止，系统正在关闭",
        "monitor": "否 - 正常系统关闭"
    },
    
    # Directory service
    8198: {
        "cause": "目录服务事件，可能是AD复制或LDAP问题",
        "monitor": "是 - 检查AD复制状态和网络连接"
    },
    
    # Network events
    16998: {
        "cause": "网络连接事件，可能是网络适配器或连接问题",
        "monitor": "视情况 - 检查网络状态"
    },
    
    # SSL/TLS errors
    36871: {
        "cause": "Schannel/SSL错误，TLS握手或证书问题",
        "monitor": "是 - 检查证书有效性和TLS配置"
    },
    36928: {
        "cause": "Schannel/SSL错误，加密通信失败",
        "monitor": "是 - 检查SSL/TLS配置和证书"
    },
    
    # High risk Error events
    4625: {
        "cause": "登录失败，可能原因：密码错误、账户锁定、暴力破解攻击",
        "monitor": "是 - 关注失败频率和来源IP"
    },
    4740: {
        "cause": "账户被锁定，通常因多次登录失败导致",
        "monitor": "是 - 检查是否为暴力破解或用户忘记密码"
    },
    4724: {
        "cause": "尝试重置用户密码，可能是管理员操作或攻击者行为",
        "monitor": "是 - 确认重置操作是否经过授权"
    },
    4726: {
        "cause": "用户账户被删除，可能是清理操作或恶意行为",
        "monitor": "是 - 确认删除操作是否经过授权"
    },
    4728: {
        "cause": "成员被添加到特权安全组（如Domain Admins）",
        "monitor": "是 - 立即确认是否为授权操作"
    },
    4732: {
        "cause": "成员被添加到本地管理员组或特权组",
        "monitor": "是 - 确认是否为授权操作"
    },
    4698: {
        "cause": "创建了计划任务，可能是持久化攻击或合法软件安装",
        "monitor": "是 - 审查任务内容和创建者"
    },
    4697: {
        "cause": "安装了新服务，可能是恶意软件持久化或合法软件安装",
        "monitor": "是 - 确认服务来源和用途"
    },
    
    # Medium risk Error events
    4662: {
        "cause": "对AD对象执行了操作，可能是正常的管理操作或目录枚举",
        "monitor": "视情况 - 关注操作频率和操作者"
    },
    4688: {
        "cause": "新进程创建，可能是正常程序运行或恶意软件执行",
        "monitor": "视情况 - 关注可疑进程名称"
    },
    4771: {
        "cause": "Kerberos预认证失败，通常是密码错误或配置问题",
        "monitor": "否 - 除非频繁发生"
    },
    4776: {
        "cause": "NTLM认证失败，可能是密码错误或旧协议使用",
        "monitor": "否 - 建议迁移到Kerberos"
    },
    
    # Warning events
    4720: {
        "cause": "创建了新用户账户，可能是新员工入职或攻击者创建后门",
        "monitor": "是 - 确认创建操作是否经过授权"
    },
    4722: {
        "cause": "用户账户被启用，可能是重新激活账户或异常操作",
        "monitor": "视情况 - 确认启用原因"
    },
    4725: {
        "cause": "用户账户被禁用，可能是员工离职或安全策略",
        "monitor": "否 - 通常是正常管理操作"
    },
    4741: {
        "cause": "创建了计算机账户，通常是新设备加入域",
        "monitor": "否 - 正常域操作"
    },
    4742: {
        "cause": "计算机账户属性变更，可能是正常的设备管理",
        "monitor": "否 - 正常域操作"
    },
    4768: {
        "cause": "Kerberos TGT票据请求，正常的域认证行为",
        "monitor": "否 - 正常认证流程"
    },
    4769: {
        "cause": "Kerberos服务票据请求，正常的资源访问行为",
        "monitor": "否 - 正常认证流程"
    },
    4778: {
        "cause": "RDP会话重新连接，用户重新连接远程桌面",
        "monitor": "否 - 正常远程操作"
    },
    5136: {
        "cause": "AD目录对象属性被修改",
        "monitor": "视情况 - 关注修改的属性和操作者"
    },
    5137: {
        "cause": "创建了AD目录对象",
        "monitor": "视情况 - 确认创建操作是否合法"
    },
    
    # System Error events
    41: {
        "cause": "系统意外关机或崩溃，可能是硬件故障、电源问题或系统错误",
        "monitor": "是 - 检查硬件健康状态和系统日志"
    },
    6008: {
        "cause": "记录了上一次意外关机事件",
        "monitor": "是 - 关注频率，可能需要硬件检查"
    },
    7031: {
        "cause": "服务意外终止，可能是服务配置错误或依赖项问题",
        "monitor": "是 - 检查服务状态和依赖关系"
    },
    7034: {
        "cause": "服务意外崩溃，可能是程序错误或资源不足",
        "monitor": "是 - 查看应用程序日志定位根因"
    },
    
    # Information events (low priority)
    4624: {
        "cause": "成功登录，正常的认证行为",
        "monitor": "否 - 正常操作"
    },
    4634: {
        "cause": "用户注销，会话结束",
        "monitor": "否 - 正常操作"
    },
    4647: {
        "cause": "用户发起注销",
        "monitor": "否 - 正常操作"
    },
    4663: {
        "cause": "尝试访问对象",
        "monitor": "否 - 正常文件访问"
    },
    4670: {
        "cause": "权限列表已更改",
        "monitor": "视情况 - 确认变更是否授权"
    },
    4702: {
        "cause": "计划任务已更新",
        "monitor": "视情况 - 确认更新内容"
    },
    4738: {
        "cause": "用户账户属性变更",
        "monitor": "否 - 正常管理操作"
    },
    4756: {
        "cause": "成员被添加到通用组",
        "monitor": "视情况 - 确认组权限级别"
    },
    4767: {
        "cause": "账户被解锁",
        "monitor": "否 - 正常管理操作"
    },
    5139: {
        "cause": "AD对象在OU之间移动",
        "monitor": "否 - 正常管理操作"
    },
    5140: {
        "cause": "访问了网络共享",
        "monitor": "否 - 正常文件访问"
    },
    5145: {
        "cause": "检查了共享对象的权限",
        "monitor": "否 - 正常文件访问"
    },
    1074: {
        "cause": "系统执行关机或重启操作",
        "monitor": "否 - 正常管理操作"
    },
    6005: {
        "cause": "事件日志服务启动，系统开始记录日志",
        "monitor": "否 - 系统启动"
    },
    6006: {
        "cause": "事件日志服务停止，系统正在关闭",
        "monitor": "否 - 系统关闭"
    },
}

def analyze_excel():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    root_dir = os.path.dirname(script_dir)
    
    input_file = os.path.join(root_dir, "AD_Log_Reports_Aggregated.xlsx")
    
    if not os.path.exists(input_file):
        print(f"文件不存在: {input_file}")
        return
    
    print(f"正在读取: {input_file}")
    
    tmp_input = os.path.join(tempfile.gettempdir(), "agg_input.xlsx")
    shutil.copy2(input_file, tmp_input)
    
    # Read sheets
    summary_df = pd.read_excel(tmp_input, sheet_name="Summary")
    events_df = pd.read_excel(tmp_input, sheet_name="EventDetails")
    
    print(f"Summary: {len(summary_df)} 行")
    print(f"EventDetails: {len(events_df)} 行")
    
    # Ensure columns exist and convert to object dtype
    if "可能原因" not in events_df.columns:
        events_df["可能原因"] = ""
    else:
        events_df["可能原因"] = events_df["可能原因"].astype(object)
    
    if "是否需要监测" not in events_df.columns:
        events_df["是否需要监测"] = ""
    else:
        events_df["是否需要监测"] = events_df["是否需要监测"].astype(object)
    
    # Process each event
    print("\n正在分析事件...")
    critical_count = 0
    error_count = 0
    warning_count = 0
    monitor_count = 0
    
    for idx, row in events_df.iterrows():
        level = str(row.get("Level", "")).strip()
        event_id = row.get("Event ID", 0)
        
        try:
            event_id = int(event_id)
        except:
            event_id = 0
        
        # Count by level
        if level == "Critical":
            critical_count += 1
        elif level == "Error":
            error_count += 1
        elif level == "Warning":
            warning_count += 1
        
        # Fill in analysis for Critical, Error, Warning
        if level in ["Critical", "Error", "Warning"]:
            if event_id in EVENT_CAUSES:
                cause = EVENT_CAUSES[event_id]["cause"]
                monitor = EVENT_CAUSES[event_id]["monitor"]
                events_df.at[idx, "可能原因"] = cause
                events_df.at[idx, "是否需要监测"] = monitor
                if monitor.startswith("是"):
                    monitor_count += 1
            else:
                events_df.at[idx, "可能原因"] = f"事件ID {event_id}，需要进一步调查"
                events_df.at[idx, "是否需要监测"] = "视情况 - 建议调查"
    
    print(f"  Critical: {critical_count}")
    print(f"  Error: {error_count}")
    print(f"  Warning: {warning_count}")
    print(f"  需要监测: {monitor_count}")
    
    # Generate overall summary
    print("\n正在生成总体总结...")
    summary_text = generate_overall_summary(summary_df, events_df, critical_count, error_count, warning_count, monitor_count)
    
    # Save to new file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(root_dir, f"AD_Log_Reports_Aggregated_{timestamp}.xlsx")
    tmp_output = os.path.join(tempfile.gettempdir(), f"agg_output_{timestamp}.xlsx")
    
    print(f"\n正在写入Excel...")
    with pd.ExcelWriter(tmp_output, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        events_df.to_excel(writer, sheet_name="EventDetails", index=False)
    
    # Add summary sheet
    wb = load_workbook(tmp_output)
    ws_summary = wb.create_sheet("总体分析", 0)
    ws_summary['A1'] = "AD日志总体分析报告"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary['A3'] = summary_text
    ws_summary['A3'].alignment = Alignment(wrap_text=True, vertical='top')
    ws_summary.column_dimensions['A'].width = 120
    
    # Apply styling
    apply_styling(wb)
    
    # Save final file
    final_path = save_file(wb, output_file)
    
    print(f"\n[OK] 完成!")
    print(f"  输出文件: {final_path}")
    print(f"  包含工作表: 总体分析, Summary, EventDetails")


def generate_overall_summary(summary_df, events_df, critical_count, error_count, warning_count, monitor_count):
    lines = []
    lines.append("=" * 80)
    lines.append("AD环境安全态势总体分析报告")
    lines.append("=" * 80)
    lines.append(f"报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"分析范围: {summary_df['AD地址'].nunique()} 台AD服务器")
    lines.append(f"总事件记录数: {len(events_df):,} 条")
    lines.append("")
    
    # Risk assessment
    lines.append("【总体风险评级】")
    if critical_count > 0:
        lines.append("  风险等级: 高风险 (Critical)")
        lines.append(f"  发现 {critical_count} 个严重事件，需要立即关注")
    elif error_count > 10:
        lines.append("  风险等级: 中高风险 (High)")
        lines.append(f"  发现 {error_count} 个错误事件，建议重点关注")
    elif warning_count > 20:
        lines.append("  风险等级: 中等风险 (Medium)")
        lines.append(f"  发现 {warning_count} 个警告事件，建议定期审查")
    else:
        lines.append("  风险等级: 低风险 (Low)")
        lines.append("  未发现明显安全风险")
    lines.append("")
    
    # Event statistics
    lines.append("【事件统计】")
    lines.append(f"  Critical (严重): {critical_count} 条")
    lines.append(f"  Error (错误): {error_count} 条")
    lines.append(f"  Warning (警告): {warning_count} 条")
    lines.append(f"  Information (信息): {len(events_df[events_df['Level'] == 'Information'])} 条")
    lines.append(f"  Log (日志): {len(events_df[events_df['Level'] == 'Log'])} 条")
    lines.append("")
    
    # Monitoring requirements
    lines.append("【监测需求】")
    lines.append(f"  需要监测的事件: {monitor_count} 条")
    lines.append("  重点关注:")
    
    # Login failures
    login_failures = events_df[events_df['Event ID'] == 4625]
    if len(login_failures) > 0:
        # Convert Count to numeric, handling any string values
        login_counts = pd.to_numeric(login_failures['Count'], errors='coerce').fillna(0)
        total_failures = int(login_counts.sum())
        lines.append(f"    - 登录失败 (EventID 4625): {total_failures:,} 次")
        if total_failures > 100:
            lines.append("      建议: 检查账户锁定策略，启用多因素认证(MFA)")
    
    # Account changes
    account_events = events_df[events_df['Event ID'].isin([4720, 4722, 4724, 4725, 4726, 4728, 4732])]
    if len(account_events) > 0:
        account_counts = pd.to_numeric(account_events['Count'], errors='coerce').fillna(0)
        total_changes = int(account_counts.sum())
        lines.append(f"    - 账户变更事件: {total_changes:,} 次")
        lines.append("      建议: 建立账户变更审批流程，定期审计特权账户")
    
    # Scheduled tasks
    task_events = events_df[events_df['Event ID'].isin([4697, 4698])]
    if len(task_events) > 0:
        task_counts = pd.to_numeric(task_events['Count'], errors='coerce').fillna(0)
        total_tasks = int(task_counts.sum())
        lines.append(f"    - 计划任务/服务变更: {total_tasks} 次")
        lines.append("      建议: 审查新建任务，确认是否为合法软件")
    
    # System issues
    system_events = events_df[events_df['Event ID'].isin([41, 6008, 7031, 7034])]
    if len(system_events) > 0:
        system_counts = pd.to_numeric(system_events['Count'], errors='coerce').fillna(0)
        total_issues = int(system_counts.sum())
        lines.append(f"    - 系统异常事件: {total_issues:,} 次")
        lines.append("      建议: 检查硬件健康状态，查看系统日志定位根因")
    
    lines.append("")
    
    # Server summary
    lines.append("【各服务器状态】")
    for _, row in summary_df.iterrows():
        ip = row.get('AD地址', '未知')
        if pd.isna(ip):
            continue
        total = row.get('总事件数', 0)
        critical = row.get('Critical', 0)
        error = row.get('Error', 0)
        warning = row.get('Warning', 0)
        
        if critical > 0:
            status = "高风险"
        elif error > 5:
            status = "中高风险"
        elif warning > 10:
            status = "中等风险"
        else:
            status = "低风险"
        
        lines.append(f"  {ip}: {status} | 总事件: {total:,} | Critical: {critical} | Error: {error} | Warning: {warning}")
    
    lines.append("")
    
    # Recommendations
    lines.append("【建议措施】")
    priority = 1
    
    if critical_count > 0:
        lines.append(f"  P{priority} [立即处理]: 调查所有严重风险事件")
        lines.append("    - 检查是否有未授权的账户操作")
        lines.append("    - 检查是否有审计日志被清除")
        priority += 1
    
    if len(login_failures) > 0 and pd.to_numeric(login_failures['Count'], errors='coerce').fillna(0).sum() > 100:
        lines.append(f"  P{priority} [高优先级]: 加强认证安全")
        lines.append("    - 审查账户锁定策略（建议5次失败锁定30分钟）")
        lines.append("    - 为特权账户启用多因素认证(MFA)")
        priority += 1
    
    if len(account_events) > 0 and pd.to_numeric(account_events['Count'], errors='coerce').fillna(0).sum() > 50:
        lines.append(f"  P{priority} [中优先级]: 完善账户管理流程")
        lines.append("    - 建立账户变更审批和审计机制")
        lines.append("    - 定期审查特权组成员（Domain Admins等）")
        priority += 1
    
    if len(task_events) > 0:
        lines.append(f"  P{priority} [定期审查]: 审查系统变更")
        lines.append("    - 审查新建的计划任务和服务")
        lines.append("    - 确认所有变更都有变更记录")
        priority += 1
    
    lines.append(f"  P{priority} [持续改进]: 建立安全基线")
    lines.append("    - 定期备份和归档日志")
    lines.append("    - 确保所有AD服务器时间同步(NTP)")
    lines.append("    - 定期执行安全评估")
    
    lines.append("")
    lines.append("=" * 80)
    lines.append("报告结束")
    lines.append("=" * 80)
    
    return "\n".join(lines)


def apply_styling(wb):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        
        # Style header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        
        # Style data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        
        # Auto-adjust column widths
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row_idx in range(1, min(ws.max_row + 1, 50)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, min(len(str(val)), 50))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def save_file(wb, output_path):
    tmp_out = os.path.join(tempfile.gettempdir(), f"final_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(tmp_out)
    
    try:
        shutil.copy2(tmp_out, output_path)
        os.remove(tmp_out)
        return output_path
    except Exception as e:
        print(f"  复制失败: {e}")
        return tmp_out


if __name__ == "__main__":
    analyze_excel()
