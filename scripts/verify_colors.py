import shutil, tempfile, os
from openpyxl import load_workbook

src = r'C:\Users\D9352\OneDrive - 基恩士（中国）有限公司\IT-PartnerShare - 文档\71. AD日志\AD_Log_Reports_Aggregated.xlsx'
tmp = os.path.join(tempfile.gettempdir(), 'verify_colors.xlsx')
shutil.copy2(src, tmp)

wb = load_workbook(tmp)
ws = wb['事件详情']

out = open(os.path.join(tempfile.gettempdir(), 'color_verify.txt'), 'w', encoding='utf-8')

# 检查前20行的颜色
out.write("前20行颜色检查:\n")
out.write("=" * 100 + "\n")

for row_idx in range(2, min(22, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row_idx, col_idx)
        header = ws.cell(1, col_idx).value
        
        if header in ['日志类型', 'AD服务器', '级别', '是否监测']:
            fill_color = cell.fill.fgColor.rgb if cell.fill.fgColor else 'None'
            value = cell.value
            row_data.append(f"{header}={value}(颜色:{fill_color})")
    
    if row_data:
        out.write(f"行{row_idx}: {' | '.join(row_data)}\n")

# 检查汇总sheet
ws_sum = wb['汇总']
out.write("\n\n汇总sheet级别列颜色:\n")
out.write("=" * 100 + "\n")

for col_idx in range(1, ws_sum.max_column + 1):
    header = ws_sum.cell(1, col_idx).value
    if header in ['Critical', 'Error', 'Warning', 'Information', 'Verbose', 'Log']:
        cell = ws_sum.cell(2, col_idx)  # 检查第二行
        fill_color = cell.fill.fgColor.rgb if cell.fill.fgColor else 'None'
        out.write(f"{header}: 颜色={fill_color}\n")

out.close()
print("[OK] 验证完成")
