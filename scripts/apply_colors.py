#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os, shutil, tempfile
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COLORS = {
    'log_type': {
        '安全': 'D6EAF8',
        '系统': 'D5F5E3',
        '应用程序': 'FDEBD0',
        '设置': 'E8DAEF',
    },
    'ad_server': {
        '10.59.91.1': 'FADBD8',
        '10.59.91.2': 'D1F2EB',
        '10.59.97.1': 'FCF3CF',
        '10.59.98.1': 'D6EAF8',
        '10.59.99.1': 'E8DAEF',
    },
    'level': {
        'Critical': 'FF0000',
        'Error': 'FF6600',
        'Warning': 'FFCC00',
        'Information': '92D050',
        'Log': '00B0F0',
        'Verbose': 'BDD7EE',
    },
    'monitor': {
        '是': 'FF9999',
        '否': 'C6EFCE',
        '视情况': 'FFEB9C',
    },
}

def get_col_index(ws, name):
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == name:
            return idx
    return None

def apply_colors():
    src = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       "AD_Log_Reports_Aggregated.xlsx")
    tmp = os.path.join(tempfile.gettempdir(), "color_input.xlsx")
    shutil.copy2(src, tmp)

    wb = load_workbook(tmp)
    ws = wb['事件详情']

    col_log = get_col_index(ws, '日志类型')
    col_server = get_col_index(ws, 'AD服务器')
    col_level = get_col_index(ws, '级别')
    col_monitor = get_col_index(ws, '是否需要监测')

    thin = Border(
        left=Side('thin'), right=Side('thin'),
        top=Side('thin'), bottom=Side('thin')
    )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # 日志类型
        if col_log:
            val = row[col_log - 1].value
            if val and str(val) in COLORS['log_type']:
                row[col_log - 1].fill = PatternFill('solid', fgColor=COLORS['log_type'][str(val)])

        # AD服务器
        if col_server:
            val = row[col_server - 1].value
            if val and str(val) in COLORS['ad_server']:
                row[col_server - 1].fill = PatternFill('solid', fgColor=COLORS['ad_server'][str(val)])

        # 级别
        if col_level:
            val = row[col_level - 1].value
            if val and str(val) in COLORS['level']:
                color = COLORS['level'][str(val)]
                row[col_level - 1].fill = PatternFill('solid', fgColor=color)
                row[col_level - 1].font = Font(bold=True, color='FFFFFF' if str(val) in ('Critical',) else '000000')

        # 是否监测
        if col_monitor:
            val = row[col_monitor - 1].value
            if val:
                val_str = str(val)
                for key, color in COLORS['monitor'].items():
                    if val_str.startswith(key):
                        row[col_monitor - 1].fill = PatternFill('solid', fgColor=color)
                        break

    # 汇总sheet也给级别列上色
    ws_sum = wb['汇总']
    # 找到Critical/Error/Warning/Information/Verbose列
    for col_idx in range(1, ws_sum.max_column + 1):
        header = ws_sum.cell(1, col_idx).value
        if header in COLORS['level']:
            color = COLORS['level'][header]
            for row_idx in range(2, ws_sum.max_row + 1):
                cell = ws_sum.cell(row_idx, col_idx)
                cell.fill = PatternFill('solid', fgColor=color)
                cell.font = Font(bold=True, color='FFFFFF' if header == 'Critical' else '000000')

    out = os.path.join(tempfile.gettempdir(), "color_output.xlsx")
    wb.save(out)
    wb.close()

    try:
        shutil.copy2(out, src)
        os.remove(tmp)
        os.remove(out)
        print("[OK] 颜色已应用")
    except Exception as e:
        print(f"覆盖失败: {e}")

if __name__ == "__main__":
    apply_colors()
