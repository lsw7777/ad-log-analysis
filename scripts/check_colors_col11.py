from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\D9352\OneDrive - 基恩士（中国）有限公司\IT-PartnerShare - 文档\71. AD日志\AD_Log_Reports_Aggregated.xlsx')
ws = wb['事件详情']

print('检查列11(是否需要监测)的颜色:')
for idx in range(120, 131):
    cell = ws.cell(idx, 11)
    value = cell.value
    color = cell.fill.fgColor.rgb if cell.fill.fgColor else 'None'
    print(f'行{idx}: 值="{value}" 颜色={color}')
