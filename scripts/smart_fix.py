#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
智能修复 AD_Log_Reports_Aggregated.xlsx
根据每个事件的具体错误消息内容生成准确的"可能原因"和"是否监测"分析
"""

import os
import shutil
import tempfile
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def analyze_event_content(event_id, provider, message):
    """
    根据Event ID、提供程序和消息内容智能分析
    返回 (可能原因, 是否监测)
    """
    if not message or message == 'nan':
        message = ""
    
    message_lower = message.lower()
    event_id = int(event_id)
    
    # 根据消息内容关键词判断
    cause = ""
    monitor = ""
    
    # 系统更新相关
    if event_id in [4, 3] and "更新" in message:
        if "重启" in message or "restart" in message_lower:
            cause = "Windows更新已下载但需要重启才能完成安装。在重启之前，系统可能缺少最新的安全补丁。"
            monitor = "是 - 建议尽快安排重启以完成更新"
        elif "未完成" in message or "pending" in message_lower:
            cause = "系统有未完成的更新事务，之前的更新或系统变更没有彻底完成。"
            monitor = "是 - 建议检查Windows Update状态并重启系统"
        else:
            cause = "Windows Update相关事件，可能涉及更新安装或配置。"
            monitor = "视情况 - 检查Windows Update状态"
    
    # 智能卡/证书相关
    elif event_id == 21 and ("智能卡" in message or "smart card" in message_lower or "证书" in message):
        cause = "智能卡登录失败，证书可能已过期、被吊销，或证书链不受信任。"
        monitor = "是 - 检查证书有效期和信任链"
    
    # IIS相关
    elif event_id == 9009 and ("IIS" in message or "web" in message_lower):
        cause = "IIS在清理历史配置目录时出错，不影响IIS正常运行。"
        monitor = "否 - 可以忽略，不影响服务"
    
    # WinRM相关
    elif event_id == 10149 and ("WinRM" in message or "远程管理" in message):
        cause = "WinRM（Windows远程管理）服务未运行，无法通过PowerShell Remoting等工具远程管理。"
        monitor = "是 - 如需远程管理，应启动WinRM服务"
    
    # DCOM权限相关
    elif event_id == 10016 and ("权限" in message or "DCOM" in message or "COM" in message):
        if "RuntimeBroker" in message:
            cause = "DCOM权限配置问题，RuntimeBroker访问被拒绝。微软官方表示通常可安全忽略。"
            monitor = "否 - 常见系统提示，可忽略"
        elif "CDP" in message or "Activity Store" in message:
            cause = "DCOM权限问题，系统组件访问被拒绝。通常不影响系统功能。"
            monitor = "否 - 可安全忽略"
        else:
            cause = "DCOM权限配置不完整，某个程序没有权限访问COM组件。"
            monitor = "视情况 - 如系统功能正常可忽略"
    
    # NTLM认证相关
    elif event_id == 6040 and ("NTLM" in message or "认证" in message):
        cause = "NTLM认证请求被拒绝，目标名称不匹配。可能是配置错误或中间人攻击。"
        monitor = "是 - 检查NTLM配置和网络安全"
    
    elif event_id == 6038 and "NTLM" in message:
        cause = "客户端使用NTLM认证（而非更安全的Kerberos）。NTLM存在已知安全漏洞。"
        monitor = "是 - 建议迁移到Kerberos认证"
    
    # 磁盘相关
    elif event_id == 32 and ("磁盘" in message or "disk" in message_lower or "写入缓存" in message):
        cause = "磁盘启用了写入缓存，突然断电可能丢失缓存数据。"
        monitor = "否 - 正常配置，确保有UPS保护"
    
    elif event_id == 32 and ("KDC" in message or "证书" in message):
        cause = "域控制器KDC证书配置不正确，缺少KDC扩展密钥用法。未加入域的设备无法通过证书认证。"
        monitor = "是 - 修复KDC证书配置"
    
    elif event_id in [58, 158] and ("磁盘" in message or "签名" in message):
        cause = "多块磁盘使用相同标识符，系统可能混淆磁盘，导致数据读写错误。"
        monitor = "是 - 立即检查并修复磁盘签名冲突"
    
    elif event_id == 157 and ("磁盘" in message or "移除" in message):
        cause = "磁盘被意外移除，可能是硬盘故障、连接松动或存储链路中断。"
        monitor = "是 - 立即检查硬件状态"
    
    elif event_id == 153 and ("磁盘" in message or "IO" in message or "重试" in message):
        cause = "磁盘进行IO重试，可能表示磁盘有坏道或连接不稳定，是硬盘故障的早期信号。"
        monitor = "是 - 建议检查磁盘健康状态"
    
    # 进程激活服务（IIS）
    elif event_id == 5840 and ("WAS" in message or "进程激活" in message):
        cause = "Windows进程激活服务以经典模式启动，IIS相关配置信息，通常不影响功能。"
        monitor = "否 - 正常配置信息"
    
    # 服务启动超时
    elif event_id == 7009 and ("启动" in message or "超时" in message):
        if "Zabbix" in message:
            cause = "Zabbix Agent启动超时，开机时资源竞争导致。"
            monitor = "否 - 常见开机现象，可忽略"
        else:
            cause = "服务启动超时，未在默认30秒内完成启动。"
            monitor = "是 - 检查服务状态和系统资源"
    
    elif event_id == 7000 and ("启动" in message or "失败" in message):
        cause = "服务启动失败，可能是账户权限不足或依赖服务未启动。"
        monitor = "是 - 检查服务配置和依赖关系"
    
    # TrustedInstaller相关
    elif event_id == 10029 and ("TrustedInstaller" in message or "等待" in message):
        cause = "系统在等待TrustedInstaller停止时超时，另一个组件因等待超时而激活失败。"
        monitor = "否 - 偶发事件，可忽略"
    
    # Netlogon信任关系
    elif event_id == 5723 and ("信任" in message or "安全通道" in message or "Netlogon" in message):
        if "找不到" in message or "not found" in message_lower:
            cause = "域控制器安全数据库中找不到计算机账户，信任关系断裂。"
            monitor = "是 - 立即检查计算机账户和AD复制"
        else:
            cause = "计算机与域控制器之间的信任关系或安全通道建立失败。"
            monitor = "是 - 检查网络连接和域信任关系"
    
    elif event_id == 5722 and ("信任" in message or "Kerberos" in message or "Netlogon" in message):
        cause = "计算机账户与域控制器之间的信任关系或Kerberos身份验证失效。"
        monitor = "是 - 检查网络连接和域控制器状态"
    
    elif event_id == 5805 and ("拒绝" in message or "访问" in message or "Netlogon" in message):
        if "信任关系" in message or "断裂" in message:
            cause = "计算机与域之间的信任关系已彻底断裂，身份验证被拒绝。"
            monitor = "是 - 立即修复计算机信任关系"
        else:
            cause = "身份验证尝试被目标服务器拒绝，可能是安全通道问题。"
            monitor = "是 - 检查网络连接和信任关系"
    
    elif event_id == 5838 and ("Netlogon" in message or "RPC" in message or "签名" in message):
        cause = "客户端使用旧版本Netlogon协议，安全级别低于域控制器要求。"
        monitor = "是 - 升级客户端Netlogon协议版本"
    
    elif event_id == 5807 and ("站点" in message or "AD站点" in message):
        cause = "客户端连接到没有映射到AD站点的域控，AD站点配置不完整，可能导致登录变慢。"
        monitor = "是 - 完善AD站点和子网配置"
    
    elif event_id == 5827 and ("Netlogon" in message or "安全" in message):
        cause = "Netlogon安全策略升级后，旧协议被拒绝。"
        monitor = "是 - 检查并升级客户端安全协议"
    
    # Kerberos相关
    elif event_id == 14 and ("Kerberos" in message or "加密" in message):
        cause = "Kerberos加密类型不匹配，请求的加密类型与账户可用加密类型不一致。"
        monitor = "是 - 检查Kerberos加密配置"
    
    # ETW相关
    elif event_id == 30 and ("ETW" in message or "权限" in message):
        cause = "ETW提供程序启用了安全机制，但事件日志服务账户缺少必要权限。"
        monitor = "是 - 修复ETW提供程序权限配置"
    
    # 注销相关
    elif event_id == 1077 and ("注销" in message or "失败" in message):
        cause = "注销操作失败，可能有程序阻止了注销。"
        monitor = "否 - 偶发事件，可忽略"
    
    # 意外关机
    elif event_id == 6008 and ("关机" in message or "重启" in message):
        cause = "系统上一次没有正常关机就重启了，可能是意外断电或系统崩溃。"
        monitor = "是 - 关注频率，可能需要硬件检查"
    
    elif event_id == 1076 and ("关机" in message or "最后" in message):
        cause = "记录谁最后一次意外关动了计算机，信息性日志。"
        monitor = "否 - 信息性日志，可忽略"
    
    # 网络通信
    elif event_id == 10028 and ("网络" in message or "DCOM" in message or "RPC" in message):
        cause = "DCOM/RPC通信链路中断，网络通信失败。"
        monitor = "是 - 检查网络连接和防火墙配置"
    
    # WMI相关
    elif event_id == 7023 and ("WMI" in message or "性能计数器" in message):
        cause = "WMI服务或性能计数器损坏，配置问题。"
        monitor = "是 - 重建WMI仓库或性能计数器"
    
    # 关机卡住
    elif event_id == 7043 and ("关机" in message or "卡住" in message):
        cause = "处理关机指令时卡住，无法按时完成关机或重启流程。"
        monitor = "是 - 检查阻止关机的进程或服务"
    
    # 服务崩溃
    elif event_id == 7034 and ("停止" in message or "崩溃" in message):
        if "安骑士" in message or "阿里云" in message:
            cause = "阿里云安骑士意外停止，安全监控功能失效。"
            monitor = "是 - 立即检查并重启安骑士服务"
        else:
            cause = "服务意外崩溃或停止。"
            monitor = "是 - 检查服务日志定位根因"
    
    elif event_id == 7031 and ("启动" in message or "失败" in message):
        cause = "服务启动时无法完成核心任务。"
        monitor = "是 - 检查服务配置和依赖关系"
    
    # 时间服务
    elif event_id == 135 and ("时间" in message or "NTP" in message):
        cause = "Windows时间服务无法连接到时间源，可能导致Kerberos认证失败。"
        monitor = "是 - 修复时间同步配置"
    
    # DNS相关
    elif event_id == 1014 and ("DNS" in message or "超时" in message):
        cause = "DNS查询超时，DNS服务器未响应。可能导致域名无法解析。"
        monitor = "是 - 检查DNS服务器和网络连接"
    
    # 组策略相关
    elif event_id == 1085 and ("组策略" in message or "MDM" in message):
        cause = "组策略中的MDM策略无法应用，设备管理策略可能未生效。"
        monitor = "是 - 检查MDM配置和网络连接"
    
    # 域加入
    elif event_id == 16998 and ("加入域" in message or "验证" in message):
        cause = "计算机加入域时安全验证失败，可能是计算机账号问题。"
        monitor = "是 - 检查计算机账户和域配置"
    
    # SSL/TLS证书
    elif event_id == 36871 and ("SSL" in message or "TLS" in message or "证书" in message or "加密" in message):
        cause = "SSL/TLS加密连接失败，可能是本地安全设置或权限问题。"
        monitor = "是 - 检查证书配置和加密策略"
    
    elif event_id == 36928 and ("OCSP" in message or "证书" in message or "吊销" in message):
        cause = "在线证书状态协议(OCSP)验证失败，无法验证证书有效性。"
        monitor = "是 - 检查OCSP服务器连接和证书配置"
    
    # 应用程序崩溃
    elif event_id == 1000 and ("崩溃" in message or "crash" in message_lower or "异常" in message):
        if "LogonUI" in message:
            cause = "登录界面程序崩溃，影响用户登录。"
            monitor = "是 - 立即检查系统文件和登录服务"
        elif "mmc" in message_lower:
            cause = "管理控制台崩溃，通常与系统文件损坏或插件冲突有关。"
            monitor = "是 - 检查管理单元和系统文件"
        elif "AzureArc" in message or "argusagent" in message:
            cause = "云监控客户端崩溃，可能是权限或兼容性问题。"
            monitor = "是 - 检查云监控服务状态"
        else:
            cause = "应用程序崩溃，需要检查具体程序的错误日志。"
            monitor = "是 - 检查应用程序日志定位根因"
    
    # 性能计数器
    elif event_id == 1008 and ("性能计数器" in message or "Perflib" in message):
        cause = "性能计数器库损坏或注册表项缺失，性能监控数据不可用。"
        monitor = "是 - 重建性能计数器库"
    
    elif event_id == 1023 and ("ntdsperf" in message_lower or "性能" in message):
        cause = "AD性能计数器DLL加载失败，通常可安全忽略。"
        monitor = "否 - 不影响AD正常运行"
    
    # 登录相关
    elif event_id == 6001 and ("Winlogon" in message or "登录" in message):
        cause = "Winlogon通知组件出错，可能影响登录/注销过程。"
        monitor = "否 - 通常不影响用户登录"
    
    elif event_id == 6004 and ("终端服务" in message or "远程桌面" in message):
        cause = "Winlogon通知终端服务组件出错，可能影响远程桌面功能。"
        monitor = "是 - 检查远程桌面服务状态"
    
    elif event_id == 6005 and ("登录" in message or "组策略" in message):
        cause = "登录过程中组件处理时间较长，可能导致登录变慢。"
        monitor = "否 - 最终会完成，可忽略"
    
    elif event_id == 6006 and ("组策略" in message or "登录" in message):
        cause = "登录过程中组策略客户端处理时间过长，拖慢登录速度。"
        monitor = "是 - 优化组策略处理"
    
    # 内存相关
    elif event_id == 2017 and ("NUMA" in message or "内存" in message):
        cause = "无法收集NUMA物理内存使用数据，不影响系统运行。"
        monitor = "否 - 不影响系统功能"
    
    # 应用程序错误
    elif event_id == 1002 and ("停止响应" in message or "崩溃" in message):
        if "mmc" in message_lower:
            cause = "管理控制台停止响应，通常与系统文件损坏或插件冲突有关。"
            monitor = "是 - 检查管理单元和系统文件"
        else:
            cause = "应用程序停止响应或崩溃。"
            monitor = "是 - 检查应用程序日志"
    
    elif event_id == 257 and ("磁盘" in message or "碎片" in message):
        cause = "磁盘碎片整理跳过某些卷，碎片太小不需要整理，磁盘状态良好。"
        monitor = "否 - 正常现象"
    
    elif event_id == 78 and ("冲突" in message or "版本" in message):
        cause = "程序与系统核心组件发生版本冲突。"
        monitor = "是 - 更新或重新安装冲突程序"
    
    # 网络相关
    elif event_id == 1014 and ("DNS" in message or "解析" in message):
        cause = "DNS查询失败，域名无法解析。"
        monitor = "是 - 检查DNS服务器配置"
    
    # 服务相关
    elif event_id == 15 and ("Defender" in message or "服务" in message):
        cause = "Microsoft Defender for Endpoint服务无法启动，缺少高级安全保护。"
        monitor = "是 - 立即检查并启动Defender服务"
    
    elif event_id == 513 and ("加密" in message or "服务" in message):
        cause = "加密服务备份系统文件失败，可能是权限不足。"
        monitor = "是 - 检查服务权限配置"
    
    # 安装相关
    elif event_id == 11722 and ("安装" in message or "Edge" in message):
        cause = "Windows Installer安装过程中某个环节失败。"
        monitor = "是 - 检查安装日志和系统环境"
    
    elif event_id == 2002 and ("EAP" in message or "DLL" in message):
        cause = "加载EAP（可扩展身份验证协议）方法时找不到DLL文件。"
        monitor = "是 - 检查身份验证配置"
    
    # 卷影复制
    elif event_id == 8193 and ("VSS" in message or "卷影" in message):
        if "关机" in message or "0x8007045b" in message:
            cause = "系统正在关机时调用VSS服务，不是故障，只是状态记录。"
            monitor = "否 - 正常关机过程"
        else:
            cause = "VSS（卷影复制服务）错误，可能影响备份功能。"
            monitor = "是 - 检查VSS服务状态"
    
    elif event_id == 12290 and ("卷影" in message or "备份" in message):
        cause = "卷影复制服务在备份过程中遇到AD数据库相关警告。"
        monitor = "是 - 检查备份完整性"
    
    elif event_id == 13 and ("VSS" in message or "关机" in message):
        cause = "系统正在关机时调用VSS服务，不是故障。"
        monitor = "否 - 正常关机过程"
    
    # 证书激活
    elif event_id == 6 and ("证书" in message or "自动注册" in message):
        cause = "自动证书注册流程在AD中找不到指定对象，与权限配置有关。"
        monitor = "是 - 检查证书模板权限"
    
    elif event_id == 8200 and ("激活" in message or "超时" in message):
        cause = "Windows激活服务器连接超时，网络连接问题。"
        monitor = "是 - 检查网络连接和激活服务器"
    
    elif event_id == 8208 and ("激活" in message or "微软" in message):
        cause = "电脑与微软激活服务器通信超时，持续的网络连接问题。"
        monitor = "是 - 检查网络连接和防火墙"
    
    elif event_id == 8198 and ("激活" in message or "许可证" in message):
        if "KMS" in message:
            cause = "无法联系KMS主机进行激活。"
            monitor = "是 - 检查KMS主机连接"
        else:
            cause = "许可证激活错误，用户登录时自动激活失败。"
            monitor = "是 - 检查激活配置和网络"
    
    elif event_id == 12293 and ("DNS" in message or "KMS" in message):
        cause = "KMS主机无法将服务记录注册到DNS区域。"
        monitor = "是 - 检查DNS记录注册权限"
    
    # Azure AD Connect
    elif event_id == 6100 and ("Azure AD" in message or "同步" in message):
        cause = "Azure AD Connect同步出问题，本地AD变更可能未同步到云端。"
        monitor = "是 - 检查Azure AD Connect同步状态"
    
    # Azure信息保护
    elif event_id == 1026 and ("MSIP" in message or "Azure" in message):
        cause = "Azure信息保护扫描程序崩溃，影响文件分类和标签功能。"
        monitor = "是 - 检查Azure信息保护服务"
    
    # 计划任务
    elif event_id == 4098 and ("计划任务" in message or "组策略" in message):
        cause = "组策略中的计划任务未成功应用。"
        monitor = "是 - 检查组策略和计划任务配置"
    
    # 用户配置
    elif event_id == 1534 and ("用户配置" in message or "服务" in message):
        cause = "用户配置文件服务通知组件出错，通常不影响用户使用。"
        monitor = "否 - 可忽略"
    
    # WMI资源
    elif event_id == 5612 and ("WMI" in message or "句柄" in message):
        cause = "WMI提供程序使用的资源超过限制，可能导致WMI查询变慢。"
        monitor = "是 - 检查WMI提供程序和监控工具"
    
    # 软件安装
    elif event_id == 119 and ("MSI" in message or "安装" in message):
        cause = "通过组策略安装软件时读取MSI文件出错，软件可能未成功安装。"
        monitor = "是 - 检查软件安装状态"
    
    # 默认处理
    else:
        if not cause:
            cause = f"事件ID {event_id}，{provider}提供程序报告的事件。"
            monitor = "视情况 - 建议查看详细日志"
    
    return cause, monitor


def fix_excel_file():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    root_dir = os.path.dirname(script_dir)
    
    input_file = os.path.join(root_dir, "AD_Log_Reports_Aggregated.xlsx")
    
    if not os.path.exists(input_file):
        print(f"文件不存在: {input_file}")
        return
    
    print(f"正在读取: {input_file}")
    
    # 复制到临时文件
    tmp_input = os.path.join(tempfile.gettempdir(), "smart_fix_input.xlsx")
    shutil.copy2(input_file, tmp_input)
    
    # 加载工作簿
    wb = load_workbook(tmp_input)
    ws = wb["事件详情"]
    
    # 找到列索引
    headers = {}
    for col_idx, cell in enumerate(ws[1], 1):
        headers[cell.value] = col_idx
    
    event_id_col = headers.get("事件ID")
    provider_col = headers.get("提供程序")
    message_col = headers.get("示例消息")  # 从示例消息工作表获取
    cause_col = headers.get("可能原因")
    monitor_col = headers.get("是否监测")
    
    if not all([event_id_col, provider_col, cause_col, monitor_col]):
        print("找不到必要的列")
        return
    
    # 读取示例消息工作表
    ws_msg = wb["示例消息"]
    msg_headers = {}
    for col_idx, cell in enumerate(ws_msg[1], 1):
        msg_headers[cell.value] = col_idx
    
    msg_event_id_col = msg_headers.get("事件ID")
    msg_content_col = msg_headers.get("示例消息")
    
    # 构建消息映射
    message_map = {}
    for row in ws_msg.iter_rows(min_row=2, values_only=True):
        if row[msg_event_id_col - 1] and row[msg_content_col - 1]:
            eid = int(row[msg_event_id_col - 1])
            message_map[eid] = str(row[msg_content_col - 1])
    
    print(f"示例消息映射: {len(message_map)} 条")
    
    # 处理每一行
    fixed_count = 0
    for row_idx in range(2, ws.max_row + 1):
        event_id = ws.cell(row_idx, event_id_col).value
        provider = ws.cell(row_idx, provider_col).value
        level = ws.cell(row_idx, headers.get("级别")).value
        
        if level not in ["Critical", "Error", "Warning"]:
            continue
        
        # 获取消息内容
        message = message_map.get(int(event_id), "")
        
        # 智能分析
        cause, monitor = analyze_event_content(event_id, provider, message)
        
        # 更新单元格
        ws.cell(row_idx, cause_col).value = cause
        ws.cell(row_idx, monitor_col).value = monitor
        
        fixed_count += 1
    
    print(f"已修复 {fixed_count} 行")
    
    # 保存
    tmp_output = os.path.join(tempfile.gettempdir(), "smart_fix_output.xlsx")
    wb.save(tmp_output)
    wb.close()
    
    # 覆盖原文件
    try:
        shutil.copy2(tmp_output, input_file)
        os.remove(tmp_input)
        os.remove(tmp_output)
        print(f"\n[OK] 修复完成!")
        print(f"已直接修改: {input_file}")
    except Exception as e:
        print(f"\n保存失败: {e}")

if __name__ == "__main__":
    fix_excel_file()
